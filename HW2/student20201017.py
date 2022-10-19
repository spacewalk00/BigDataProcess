import openpyxl

wb = openpyxl.load_workbook( "student.xlsx" )
ws = wb['Sheet1']

row_id = 1
student_num = 0
total = [] 
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total.append(sum_v)
		student_num += 1	
	row_id += 1

# grade
total_sorted = [] 
total_sorted = sorted(total)
total_sorted.reverse()
#print(total_sorted)

rank = []
dic = {}

for t in total:
	rank.append(total_sorted.index(t)+1)

	if total_sorted.index(t)+1 in dic:
		dic[total_sorted.index(t)+1] += 1
	else:
		dic[total_sorted.index(t)+1] = 1
#print(dic)
#print(rank)

i = 0
for r in rank:
	rank[i] += dic[rank[i]]	- 1	
	i += 1
	
#print(rank)

grade_list = []
abc = [ [], [], [] ]


for r in rank:
	a_condition = student_num * 0.3
	b_condition = student_num * 0.7
	
	if a_condition >= r and len(abc[0]) < a_condition:
		grade = "A"
		abc[0].append(r)
	elif b_condition  >= r and len(abc[1]) + len(abc[0]) < b_condition:
		grade = "B"
		abc[1].append(r)
	else:
		grade = "C"
		abc[2].append(r)

	grade_list.append(grade)	
		
#print(abc)
#print(grade_list)


start_index = 0
for a in range(3):
	if a == 1:
		start_index = len(abc[0])
	elif a == 2:
		start_index = len(abc[1]) + len(abc[0])

	#print('sorted',  sorted_abc )
	index = 0
	
	#키값이 엑셀순 index, 값이 rank인 dic2을 만들기
	dic2 = {}
	for d in range(len(abc[a])):
		dic2[d] = abc[a][d]
	#print(dic2)
	#dic2 정렬
	sorted_dic2 = dict(sorted(dic2.items(), key=lambda x: x[1]))
	#print(sorted_dic2)
	#요소가 tuple인 list 만들기 , 정렬된 아이들 [ (엑셀 index, rank), .. ] sorted_tuple[0][0]
	sorted_tuple = sorted(dic2.items(), key=lambda x: x[1])
	#print(sorted_tuple)
	
	for b in range(len(abc[a])):
		#print(sorted_tuple[b][0] + start_index)
		#print('맨앞', sorted_tuple[0][1])
		#print('몇개있냐', rank.count(sorted_tuple[b][1]))
		if len(abc[a])/2 > rank.count(sorted_tuple[b][1]):
			grade_list[sorted_tuple[b][0] + start_index] += '+'
		else:
			grade_list[sorted_tuple[b][0] + start_index] += '0'
#print(grade_list)

row_id = 1 
for row in ws:
	if row_id != 1:	
		ws.cell(row = row_id, column = 8).value = grade_list.pop(0) 
	row_id += 1

wb.save("student.xlsx")
